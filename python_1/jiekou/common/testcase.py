# _*_ coding:utf-8 _*_
# @Time    :2019/6/17 13:41
# @Author  :xiuhong.guo
# @Email   :892336120@qq.com
# @File    :testcase.py

# 代码框架结构：测试数据和脚本分离，方便管理和维护
# 一般分为测试数据包（如data）、测试逻辑封装包（如common）、测试用例包（如testcases）、结果报表包（如report）、
# 日志包（如log）、配置文件包（如config）

# 读取Excel表格中的接口测试用例数据
import openpyxl
from jiekou.common import request_home_1 # 引入发起requests请求的模块
from jiekou.common import contants # 引入获取文件路径的模块

class Case:
    '''测试用例类，每个测试用例实际上就是它的一个实例'''
    def __init__(self):
        self.case_id=None
        self.title=None
        self.url=None
        self.data = None
        self.method = None
        self.expected = None
        self.actual = None
        self.result = None
class DoExcel:
    def __init__(self,file_name,sheet_name):
        # 异常处理，当打开的文件不存在
        try:
            self.file_name=file_name
            self.sheet_name=sheet_name
            self.workbook=openpyxl.load_workbook(file_name)
            self.sheet=self.workbook[sheet_name]
        except:
            print('文件不存在')
    def get_cases(self):
        max_row=self.sheet.max_row # 获取最大行
        cases=[] # 存放所有的测试用例
        for r in range(2,max_row+1):
            case=Case() # Case实例，可以把测试数据以属性的形式存放在Case对象中，类似于字典，有键值对
            case.case_id = self.sheet.cell(row=r,column=1).value
            case.title = self.sheet.cell(row=r, column=2).value
            case.url = self.sheet.cell(row=r,column=3).value
            case.data = self.sheet.cell(row=r, column=4).value
            case.method= self.sheet.cell(row=r,column=5).value
            case.expected = self.sheet.cell(row=r, column=6).value
            cases.append(case)
            self.workbook.close()
        return cases # 返回cases列表，cases列表里有多个case对象，case对象里有很多属性
    def write_result(self,row,actual,result):
        sheet=self.workbook[self.sheet_name]
        sheet.cell(row,7).value=actual
        sheet.cell(row,8).value = result
        self.workbook.save(self.file_name)
        self.workbook.close()

if __name__ == '__main__':
    # contants.case_file是引入的contants模块中获取的表格地址
    do_excel=DoExcel(contants.case_file,sheet_name='login')
    cases=do_excel.get_cases()
    http_request= request_home_1.HTTPRequest()
    for case in cases: # case是个对象，存放在cases列表中
        # print(case.title)
        # print(case.method)
        print(case.__dict__) # 每个对象都有一个dict属性，用来以字典的形式显示每一个对象中属性的数据
        resp=http_request.request(case.method,case.url,case.data) # 接口测试中的参数data要传字典，
        # 然而从表格读出来的data是个字符串，这里在接口测试的home_1模块中进行了data数据类型的转换
        actual=resp.text
        print(resp.text) # {"status":1,"code":"10001","data":null,"msg":"登录成功"}
        resp_dict=resp.json()  # 把字符串转换成字典，就可以取响应文本中具体某个字段的值
        print(resp_dict) # {'msg': '登录成功', 'data': None, 'status': 1, 'code': '10001'}，
        # 转换成字典后，null就会变成None
        if case.expected == actual: # 判断期望结果是否与实际结果一致
            do_excel.write_result(case.case_id + 1,actual,'PASS')
        else:
            do_excel.write_result(case.case_id + 1,actual,'FAIL')